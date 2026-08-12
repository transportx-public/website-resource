#!/usr/bin/env python3
"""Generate one Hugo group-meeting event per date from literature sharing records."""

from __future__ import annotations

import argparse
import json
import re
import shutil
from collections import defaultdict
from datetime import date, datetime
from pathlib import Path

from openpyxl import load_workbook


REPO_ROOT = Path(__file__).resolve().parents[1]
DEFAULT_INPUT = REPO_ROOT / "data-input" / "events" / "events.xlsx"
DEFAULT_OUTPUT = REPO_ROOT / "content" / "event"

PROFILE_BY_READER = {
    "卢永晟": "Yongsheng Lu",
    "黄知微": "Zhiwei Huang",
    "李毅喆": "Yizhe Li",
    "夏强": "Qiang Xia",
    "果金杉": "Jinshan Guo",
    "董昱菡": "Yuhan Dong",
    "Abdi": "IBRAHIM ABDI MOHAMMED",
}


def text(value: object) -> str:
    if value is None:
        return ""
    return str(value).strip()


def visible_text(value: object) -> str:
    return text(value).replace("—", "-").replace("–", "-")


def yaml_scalar(value: object) -> str:
    if isinstance(value, bool):
        return "true" if value else "false"
    if value is None:
        return "''"
    return json.dumps(str(value), ensure_ascii=False)


def parse_date(value: object, field_name: str) -> date:
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value

    raw = text(value)
    for fmt in ("%Y-%m-%d", "%Y/%m/%d"):
        try:
            return datetime.strptime(raw, fmt).date()
        except ValueError:
            pass
    raise ValueError(f"Invalid {field_name}: {raw}")


def publication_month(value: object) -> str:
    if value in (None, ""):
        return ""
    if isinstance(value, datetime):
        return value.strftime("%Y-%m")
    if isinstance(value, date):
        return value.strftime("%Y-%m")

    raw = text(value)
    match = re.match(r"^(\d{4})[-/年](\d{1,2})", raw)
    if not match:
        raise ValueError(f"Invalid 发布日期: {raw}")
    return f"{int(match.group(1)):04d}-{int(match.group(2)):02d}"


def first_value(row: dict[str, object], *names: str) -> str:
    for name in names:
        value = text(row.get(name))
        if value:
            return value
    return ""


def read_rows(path: Path) -> list[dict[str, object]]:
    workbook = load_workbook(path, data_only=True)
    sheet = workbook[workbook.sheetnames[0]]
    values = list(sheet.iter_rows(values_only=True))
    if not values:
        return []

    headers = [text(cell) for cell in values[0]]
    required = {"名称", "序号", "汇报人", "日期", "发布日期", "出处", "亮点"}
    missing = required.difference(headers)
    if missing:
        raise ValueError(f"events.xlsx is missing columns: {', '.join(sorted(missing))}")

    rows: list[dict[str, object]] = []
    for values_row in values[1:]:
        if not any(value not in (None, "") for value in values_row):
            continue
        row = dict(zip(headers, values_row))
        title = visible_text(row.get("名称"))
        reader = text(row.get("汇报人"))
        if not title or not reader or row.get("日期") in (None, ""):
            raise ValueError(f"文献记录缺少名称、汇报人或日期: {row}")
        if reader not in PROFILE_BY_READER:
            raise ValueError(f"No People profile mapping for reader: {reader}")
        rows.append(row)
    return rows


def render_member(lines: list[str], member: dict[str, str], indent: str) -> None:
    lines.extend(
        [
            f"{indent}- name: {yaml_scalar(member['name'])}",
            f"{indent}  profile: {yaml_scalar(member['profile'])}",
        ]
    )


def render_reading(lines: list[str], row: dict[str, object]) -> None:
    reader = text(row.get("汇报人"))
    sequence = int(row["序号"])
    lines.extend(
        [
            f"  - title: {yaml_scalar(visible_text(row.get('名称')))}",
            f"    sequence: {sequence}",
            f"    reader_name: {yaml_scalar(reader)}",
            f"    reader_profile: {yaml_scalar(PROFILE_BY_READER[reader])}",
            f"    published: {yaml_scalar(publication_month(row.get('发布日期')))}",
            f"    source: {yaml_scalar(visible_text(row.get('出处')))}",
            f"    highlight: {yaml_scalar(visible_text(row.get('亮点')))}",
            f"    url: {yaml_scalar(first_value(row, '原文链接', 'paper_url'))}",
            f"    slides: {yaml_scalar(first_value(row, '文献分享PPT', 'reading_slides'))}",
        ]
    )


def render_event(meeting_date: date, rows: list[dict[str, object]]) -> tuple[str, str]:
    rows.sort(key=lambda row: int(row["序号"]))
    day = meeting_date.strftime("%Y-%m-%d")
    folder = f"group-meeting-{day}"
    summary = f"本次组会收录 {len(rows)} 篇文献阅读分享。"

    members: list[dict[str, str]] = []
    seen_members: set[str] = set()
    for row in rows:
        reader = text(row.get("汇报人"))
        if reader not in seen_members:
            members.append({"name": reader, "profile": PROFILE_BY_READER[reader]})
            seen_members.add(reader)

    meeting_slides = next(
        (first_value(row, "组会PPT", "meeting_slides") for row in rows if first_value(row, "组会PPT", "meeting_slides")),
        "",
    )
    meeting_notes = next(
        (first_value(row, "会议纪要", "meeting_notes") for row in rows if first_value(row, "会议纪要", "meeting_notes")),
        "",
    )

    lines = [
        "---",
        f"title: {yaml_scalar(f'组会 | {day}')}",
        f"event: {yaml_scalar('课题组组会')}",
        f"summary: {yaml_scalar(summary)}",
        "abstract: ''",
        f"date: {yaml_scalar(day + 'T00:00:00+08:00')}",
        "all_day: true",
        f"publishDate: {yaml_scalar(day + 'T00:00:00+08:00')}",
        "authors: []",
        "tags:",
        f"  - {yaml_scalar('组会')}",
        f"  - {yaml_scalar('文献分享')}",
        "featured: false",
        "image:",
        "  caption: ''",
        "  focal_point: Center",
        "url_code: ''",
        "url_pdf: ''",
        "url_slides: ''",
        "url_video: ''",
        "slides: ''",
        "projects: []",
        f"meeting_slides: {yaml_scalar(meeting_slides)}",
        f"meeting_notes: {yaml_scalar(meeting_notes)}",
        "reading_members:",
    ]
    for member in members:
        render_member(lines, member, "  ")

    lines.append("readings:")
    for row in rows:
        render_reading(lines, row)
    lines.extend(["---", ""])
    return folder, "\n".join(lines)


def generate(input_path: Path, output_dir: Path) -> tuple[int, int]:
    rows = read_rows(input_path)
    meetings: dict[date, list[dict[str, object]]] = defaultdict(list)
    for row in rows:
        meetings[parse_date(row["日期"], "日期")].append(row)

    output_dir.mkdir(parents=True, exist_ok=True)
    for stale_dir in output_dir.glob("group-meeting-*"):
        if stale_dir.is_dir():
            shutil.rmtree(stale_dir)

    for meeting_date in sorted(meetings):
        folder, content = render_event(meeting_date, meetings[meeting_date])
        event_dir = output_dir / folder
        event_dir.mkdir(parents=True, exist_ok=True)
        (event_dir / "index.md").write_text(content, encoding="utf-8")

    print(f"Generated {len(meetings)} group meetings from {len(rows)} literature records")
    return len(meetings), len(rows)


def main() -> None:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--input", type=Path, default=DEFAULT_INPUT)
    parser.add_argument("--output", type=Path, default=DEFAULT_OUTPUT)
    args = parser.parse_args()
    generate(args.input, args.output)


if __name__ == "__main__":
    main()
