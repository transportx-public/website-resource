#!/usr/bin/env python3
"""Generate Hugo author profiles from the TransportX people workbook."""

from __future__ import annotations

import argparse
import json
import re
import shutil
from dataclasses import dataclass
from pathlib import Path
from typing import Iterable

from openpyxl import load_workbook


REPO_ROOT = Path(__file__).resolve().parents[1]
DEFAULT_EXCEL = REPO_ROOT.parent / "附件下载_People清单" / "people.xlsx"
DEFAULT_AVATARS = REPO_ROOT.parent / "附件下载_People清单" / "avatars"
DEFAULT_AUTHORS = REPO_ROOT / "content" / "authors"


@dataclass(frozen=True)
class Person:
    name: str
    name_chinese: str
    picture: str
    degree: str
    status: str
    enrollment_year: int | None
    introduction: str
    interests: list[str]
    email: str
    website: str
    organization: str
    other_information: str


def text(value: object) -> str:
    if value is None:
        return ""
    return str(value).strip()


def yaml_scalar(value: object) -> str:
    if isinstance(value, bool):
        return "true" if value else "false"
    if value is None:
        return "''"
    return json.dumps(str(value), ensure_ascii=False)


def split_name(name: str) -> tuple[str, str]:
    parts = name.split()
    if not parts:
        return "", ""
    if len(parts) == 1:
        return parts[0], ""
    return " ".join(parts[:-1]), parts[-1]


def split_interests(value: str) -> list[str]:
    items = [part.strip() for part in re.split(r"[,，;；]", value) if part.strip()]
    return items


def safe_folder_name(name: str) -> str:
    cleaned = re.sub(r"[/\\:*?\"<>|]", "-", name).strip()
    return re.sub(r"\s+", " ", cleaned)


def role_label(degree: str, status: str) -> str:
    degree = " ".join(degree.replace("PhD", "Ph.D").split())
    status = " ".join(status.split())
    if not degree:
        return status or "People"
    if not status:
        return degree
    return f"{degree} {status}"


def year_group(person: Person) -> str:
    if person.enrollment_year is None:
        return "Unknown Year"
    return str(person.enrollment_year)


def read_people(path: Path) -> list[Person]:
    workbook = load_workbook(path, data_only=True)
    sheet = workbook[workbook.sheetnames[0]]
    rows = list(sheet.iter_rows(values_only=True))
    if not rows:
        return []

    headers = [text(cell) for cell in rows[0]]
    required = {
        "Name",
        "Name_Chinese",
        "picture",
        "user_groups",
        "is_gradiate",
        "enrollment_year",
        "introduction",
        "Interests",
        "email",
        "Website",
        "organizations",
        "other information",
    }
    missing = sorted(required.difference(headers))
    if missing:
        raise ValueError(f"Missing required columns in {path}: {', '.join(missing)}")

    people: list[Person] = []
    for row_number, row in enumerate(rows[1:], start=2):
        data = dict(zip(headers, row))
        name = text(data.get("Name"))
        if not name:
            continue
        year_value = data.get("enrollment_year")
        try:
            enrollment_year = int(year_value) if year_value not in (None, "") else None
        except (TypeError, ValueError) as exc:
            raise ValueError(f"Invalid enrollment_year on row {row_number}: {year_value}") from exc

        people.append(
            Person(
                name=name,
                name_chinese=text(data.get("Name_Chinese")),
                picture=text(data.get("picture")),
                degree=text(data.get("user_groups")),
                status=text(data.get("is_gradiate")),
                enrollment_year=enrollment_year,
                introduction=text(data.get("introduction")),
                interests=split_interests(text(data.get("Interests"))),
                email=text(data.get("email")),
                website=text(data.get("Website")),
                organization=text(data.get("organizations")),
                other_information=text(data.get("other information")),
            )
        )
    return people


def avatar_index(avatars_dir: Path) -> dict[str, Path]:
    avatars: dict[str, Path] = {}
    for path in avatars_dir.iterdir():
        if path.is_file():
            avatars[path.name.lower()] = path
    return avatars


def find_avatar(person: Person, avatars: dict[str, Path]) -> Path | None:
    if person.picture:
        match = avatars.get(person.picture.lower())
        if match:
            return match
    expected_stem = re.sub(r"\s+", "_", person.name.lower())
    for path in avatars.values():
        if path.stem.lower() == expected_stem:
            return path
    return None


def yaml_list(key: str, values: Iterable[str]) -> list[str]:
    values = [value for value in values if value]
    if not values:
        return [f"{key}: []"]
    lines = [f"{key}:"]
    lines.extend(f"  - {yaml_scalar(value)}" for value in values)
    return lines


def render_author(person: Person) -> str:
    first_name, last_name = split_name(person.name)
    lines = [
        "---",
        f"title: {yaml_scalar(person.name)}",
        f"first_name: {yaml_scalar(first_name)}",
        f"last_name: {yaml_scalar(last_name)}",
        "authors:",
        f"  - {yaml_scalar(person.name)}",
        "superuser: false",
        f"role: {yaml_scalar(role_label(person.degree, person.status))}",
    ]

    if person.organization:
        lines.extend(
            [
                "organizations:",
                f"  - name: {yaml_scalar(person.organization)}",
                "    url: ''",
            ]
        )
    else:
        lines.append("organizations: []")

    lines.extend(
        [
            f"bio: {yaml_scalar(person.introduction)}",
            *yaml_list("interests", person.interests),
        ]
    )
    if person.enrollment_year is not None:
        lines.append(f"enrollment_year: {person.enrollment_year}")
    if person.website:
        lines.append(f"website: {yaml_scalar(person.website)}")
    if person.picture:
        lines.append(f"avatar_source: {yaml_scalar(person.picture)}")
    if person.degree:
        lines.append(f"academic_degree: {yaml_scalar(person.degree)}")
    if person.status:
        lines.append(f"academic_status: {yaml_scalar(person.status)}")
    if person.name_chinese:
        lines.append(f"name_chinese: {yaml_scalar(person.name_chinese)}")
    if person.other_information:
        lines.append(f"other_information: {yaml_scalar(person.other_information)}")

    lines.extend(
        [
            "education:",
            "  courses: []",
            f"email: {yaml_scalar(person.email)}",
        ]
    )

    social: list[tuple[str, str, str]] = []
    if person.email:
        social.append(("envelope", "fas", f"mailto:{person.email}"))
    if person.website:
        social.append(("globe", "fas", person.website))
    if social:
        lines.append("social:")
        for icon, icon_pack, link in social:
            lines.extend(
                [
                    f"  - icon: {yaml_scalar(icon)}",
                    f"    icon_pack: {yaml_scalar(icon_pack)}",
                    f"    link: {yaml_scalar(link)}",
                ]
            )
    else:
        lines.append("social: []")

    lines.extend(
        [
            "highlight_name: false",
            "user_groups:",
            f"  - {yaml_scalar(year_group(person))}",
            "---",
            "",
        ]
    )

    if person.introduction:
        lines.extend([person.introduction, ""])
    return "\n".join(lines)


def generate(
    people: list[Person],
    avatars_dir: Path,
    authors_dir: Path,
    dry_run: bool,
) -> list[str]:
    avatars = avatar_index(avatars_dir)
    warnings: list[str] = []
    planned: list[tuple[Person, Path | None]] = []

    seen_folders: set[str] = set()
    for person in people:
        folder = safe_folder_name(person.name)
        if folder in seen_folders:
            raise ValueError(f"Duplicate author folder name: {folder}")
        seen_folders.add(folder)
        avatar = find_avatar(person, avatars)
        if avatar is None:
            warnings.append(f"Missing avatar for {person.name}: {person.picture or '(blank)'}")
        planned.append((person, avatar))

    if dry_run:
        return warnings

    authors_dir.mkdir(parents=True, exist_ok=True)
    for child in authors_dir.iterdir():
        if child.is_dir():
            shutil.rmtree(child)
        else:
            child.unlink()

    for person, avatar in planned:
        person_dir = authors_dir / safe_folder_name(person.name)
        person_dir.mkdir(parents=True, exist_ok=True)
        (person_dir / "_index.md").write_text(render_author(person), encoding="utf-8")
        if avatar is not None:
            shutil.copy2(avatar, person_dir / f"avatar{avatar.suffix.lower()}")

    return warnings


def main() -> None:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--excel", type=Path, default=DEFAULT_EXCEL)
    parser.add_argument("--avatars-dir", type=Path, default=DEFAULT_AVATARS)
    parser.add_argument("--authors-dir", type=Path, default=DEFAULT_AUTHORS)
    parser.add_argument("--dry-run", action="store_true")
    args = parser.parse_args()

    people = read_people(args.excel)
    warnings = generate(people, args.avatars_dir, args.authors_dir, args.dry_run)

    action = "Would generate" if args.dry_run else "Generated"
    print(f"{action} {len(people)} author profiles in {args.authors_dir}")
    if warnings:
        print("Warnings:")
        for warning in warnings:
            print(f"- {warning}")


if __name__ == "__main__":
    main()
