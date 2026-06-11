#!/usr/bin/env python3
"""Generate Hugo news posts from data-input/posts/posts.xlsx."""

from __future__ import annotations

import argparse
import json
import re
import shutil
from datetime import date, datetime, time
from pathlib import Path

from openpyxl import load_workbook


REPO_ROOT = Path(__file__).resolve().parents[1]
DEFAULT_INPUT = REPO_ROOT / "data-input" / "posts" / "posts.xlsx"
DEFAULT_PICTURES = REPO_ROOT / "data-input" / "posts" / "pictures"
DEFAULT_MARKDOWN = REPO_ROOT / "data-input" / "posts" / "markdown"
DEFAULT_OUTPUT = REPO_ROOT / "content" / "post"


def text(value: object) -> str:
    if value is None:
        return ""
    return str(value).strip()


def yaml_scalar(value: object) -> str:
    if isinstance(value, bool):
        return "true" if value else "false"
    if value is None:
        return "''"
    return json.dumps(str(value), ensure_ascii=False)


def split_list(value: str) -> list[str]:
    return [part.strip() for part in re.split(r"[,，;；]", value) if part.strip()]


def parse_bool(value: object, default: bool = False) -> bool:
    raw = text(value).lower()
    if not raw:
        return default
    return raw in {"1", "true", "yes", "y", "是", "发布"}


def parse_date(value: object) -> datetime | None:
    if value in (None, ""):
        return None
    if isinstance(value, datetime):
        return value
    if isinstance(value, date):
        return datetime.combine(value, time.min)
    raw = text(value)
    for fmt in ("%Y-%m-%d", "%Y/%m/%d", "%Y-%m-%d %H:%M", "%Y/%m/%d %H:%M"):
        try:
            return datetime.strptime(raw, fmt)
        except ValueError:
            pass
    try:
        return datetime.fromisoformat(raw)
    except ValueError as exc:
        raise ValueError(f"Invalid post date value: {raw}") from exc


def format_date(value: datetime | None) -> str:
    if value is None:
        return "1970-01-01"
    return value.strftime("%Y-%m-%d")


def slugify(value: str, fallback: str) -> str:
    slug = re.sub(r"[^0-9A-Za-z\u4e00-\u9fff]+", "-", value).strip("-").lower()
    return slug or fallback


def read_rows(path: Path) -> list[dict[str, object]]:
    workbook = load_workbook(path, data_only=True)
    sheet = workbook[workbook.sheetnames[0]]
    rows = list(sheet.iter_rows(values_only=True))
    if not rows:
        return []
    headers = [text(cell) for cell in rows[0]]
    if "title" not in headers or "date" not in headers:
        raise ValueError("posts.xlsx must include at least title and date columns")
    return [dict(zip(headers, row)) for row in rows[1:]]


def strip_front_matter(markdown: str) -> str:
    if not markdown.startswith("---"):
        return markdown.strip()
    parts = markdown.split("---", 2)
    if len(parts) < 3:
        return markdown.strip()
    return parts[2].strip()


def body_from_markdown(row: dict[str, object], markdown_dir: Path) -> str:
    body = text(row.get("body"))
    markdown_file = text(row.get("markdown_file"))
    if markdown_file:
        source = markdown_dir / markdown_file
        if source.is_file():
            return strip_front_matter(source.read_text(encoding="utf-8"))
        print(f"WARNING: missing post markdown file: {source}")
    return body


def copy_picture(row: dict[str, object], post_dir: Path, pictures_dir: Path) -> None:
    picture = text(row.get("picture"))
    if not picture:
        return
    source = pictures_dir / picture
    if not source.is_file():
        print(f"WARNING: missing post picture: {source}")
        return
    shutil.copyfile(source, post_dir / f"featured{source.suffix.lower()}")


def render_post(row: dict[str, object], markdown_dir: Path) -> tuple[str, str] | None:
    if "publish" in row and not parse_bool(row.get("publish")):
        return None

    title = text(row.get("title"))
    if not title:
        return None
    published = parse_date(row.get("date"))
    if published is None:
        raise ValueError(f"Missing date for post: {title}")
    day = format_date(published)
    folder = slugify(text(row.get("slug")) or f"{day}-{title}", f"post-{published.strftime('%Y%m%d')}")
    tags = split_list(text(row.get("tags")))
    authors = split_list(text(row.get("authors")))

    lines = [
        "---",
        f"title: {yaml_scalar(title)}",
        f"date: {yaml_scalar(day)}",
        f"summary: {yaml_scalar(text(row.get('summary')))}",
    ]
    if authors:
        lines.append("authors:")
        lines.extend(f"  - {yaml_scalar(author)}" for author in authors)
    else:
        lines.append("authors: []")
    if tags:
        lines.append("tags:")
        lines.extend(f"  - {yaml_scalar(tag)}" for tag in tags)
    else:
        lines.append("tags: []")
    lines.extend(
        [
            f"featured: {yaml_scalar(parse_bool(row.get('featured')))}",
            "image:",
            f"  caption: {yaml_scalar(text(row.get('caption')))}",
            "  focal_point: top",
            "---",
            "",
        ]
    )
    body = body_from_markdown(row, markdown_dir)
    if body:
        lines.extend([body, ""])
    return folder, "\n".join(lines)


def generate(input_path: Path, output_dir: Path, pictures_dir: Path, markdown_dir: Path) -> int:
    rows = read_rows(input_path)
    count = 0
    for row in rows:
        rendered = render_post(row, markdown_dir)
        if rendered is None:
            continue
        folder, content = rendered
        post_dir = output_dir / folder
        post_dir.mkdir(parents=True, exist_ok=True)
        (post_dir / "index.md").write_text(content, encoding="utf-8")
        copy_picture(row, post_dir, pictures_dir)
        count += 1
    print(f"Generated {count} post pages in {output_dir}")
    return count


def main() -> None:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--input", type=Path, default=DEFAULT_INPUT)
    parser.add_argument("--pictures", type=Path, default=DEFAULT_PICTURES)
    parser.add_argument("--markdown", type=Path, default=DEFAULT_MARKDOWN)
    parser.add_argument("--output", type=Path, default=DEFAULT_OUTPUT)
    args = parser.parse_args()
    generate(args.input, args.output, args.pictures, args.markdown)


if __name__ == "__main__":
    main()
