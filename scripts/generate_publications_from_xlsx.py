#!/usr/bin/env python3
"""Generate Hugo publication pages from data-input/publications/publications.xlsx."""

from __future__ import annotations

import argparse
import json
import re
import shutil
from datetime import date, datetime, time
from pathlib import Path

from openpyxl import load_workbook


REPO_ROOT = Path(__file__).resolve().parents[1]
DEFAULT_INPUT = REPO_ROOT / "data-input" / "publications" / "publications.xlsx"
DEFAULT_PICTURES = REPO_ROOT / "data-input" / "publications" / "pictures"
DEFAULT_MARKDOWN = REPO_ROOT / "data-input" / "publications" / "markdown"
DEFAULT_PDFS = REPO_ROOT / "data-input" / "publications" / "pdfs"
DEFAULT_OUTPUT = REPO_ROOT / "content" / "publication"


def text(value: object) -> str:
    if value is None:
        return ""
    return str(value).strip()


def yaml_scalar(value: object) -> str:
    if isinstance(value, bool):
        return "true" if value else "false"
    if value is None:
        return "''"
    return json.dumps(str(value), ensure_ascii=False)


def split_list(value: str) -> list[str]:
    return [part.strip() for part in re.split(r"[,，;；]", value) if part.strip()]


def parse_bool(value: object, default: bool = False) -> bool:
    raw = text(value).lower()
    if not raw:
        return default
    return raw in {"1", "true", "yes", "y", "是", "发布"}


def parse_date(value: object) -> datetime | None:
    if value in (None, ""):
        return None
    if isinstance(value, datetime):
        return value
    if isinstance(value, date):
        return datetime.combine(value, time.min)
    raw = text(value)
    for fmt in ("%Y-%m-%d", "%Y/%m/%d", "%Y-%m-%d %H:%M", "%Y/%m/%d %H:%M"):
        try:
            return datetime.strptime(raw, fmt)
        except ValueError:
            pass
    try:
        return datetime.fromisoformat(raw)
    except ValueError as exc:
        raise ValueError(f"Invalid publication date value: {raw}") from exc


def format_date(value: datetime | None) -> str:
    if value is None:
        return "1970-01-01T00:00:00+08:00"
    return value.strftime("%Y-%m-%dT%H:%M:%S+08:00")


def slugify(value: str, fallback: str) -> str:
    slug = re.sub(r"[^0-9A-Za-z\u4e00-\u9fff]+", "-", value).strip("-").lower()
    return slug or fallback


def read_rows(path: Path) -> list[dict[str, object]]:
    workbook = load_workbook(path, data_only=True)
    sheet = workbook[workbook.sheetnames[0]]
    rows = list(sheet.iter_rows(values_only=True))
    if not rows:
        return []
    headers = [text(cell) for cell in rows[0]]
    required = {"title", "authors", "date", "publication_types", "publication", "abstract", "tags"}
    missing = sorted(required.difference(headers))
    if missing:
        raise ValueError(f"publications.xlsx missing columns: {', '.join(missing)}")
    return [dict(zip(headers, row)) for row in rows[1:]]


def copy_optional_file(filename: str, source_dir: Path, target_dir: Path, target_name: str | None = None) -> str:
    if not filename:
        return ""
    source = source_dir / filename
    if not source.is_file():
        print(f"WARNING: missing publication file: {source}")
        return ""
    name = target_name or source.name
    shutil.copyfile(source, target_dir / name)
    return name


def body_from_markdown(row: dict[str, object], markdown_dir: Path) -> str:
    body = text(row.get("body"))
    markdown_file = text(row.get("markdown_file"))
    if markdown_file:
        source = markdown_dir / markdown_file
        if source.is_file():
            return source.read_text(encoding="utf-8").strip()
        print(f"WARNING: missing publication markdown file: {source}")
    return body


def render_publication(row: dict[str, object], markdown_dir: Path) -> tuple[str, str] | None:
    if "publish" in row and not parse_bool(row.get("publish"), default=True):
        return None

    title = text(row.get("title"))
    if not title:
        return None
    published = parse_date(row.get("date"))
    day = published.strftime("%Y-%m-%d") if published else "1970-01-01"
    folder = slugify(text(row.get("slug")) or f"{day}-{title}", f"publication-{day}")
    authors = split_list(text(row.get("authors")))
    tags = split_list(text(row.get("tags")))
    publication_types = split_list(text(row.get("publication_types"))) or [text(row.get("publication_types"))]

    pdf_file = text(row.get("pdf_file"))
    url_pdf = text(row.get("url_pdf"))
    if pdf_file and not url_pdf:
        url_pdf = pdf_file

    lines = [
        "---",
        f"title: {yaml_scalar(title)}",
    ]
    if authors:
        lines.append("authors:")
        lines.extend(f"  - {yaml_scalar(author)}" for author in authors)
    else:
        lines.append("authors: []")
    lines.extend(
        [
            f"date: {yaml_scalar(format_date(published))}",
            f"doi: {yaml_scalar(text(row.get('doi')))}",
            f"publishDate: {yaml_scalar(day + 'T00:00:00+08:00')}",
        ]
    )
    publication_types = [item for item in publication_types if item]
    if publication_types:
        lines.append("publication_types:")
        lines.extend(f"  - {yaml_scalar(item)}" for item in publication_types)
    else:
        lines.append("publication_types: []")
    lines.extend(
        [
            f"publication: {yaml_scalar(text(row.get('publication')))}",
            f"publication_short: {yaml_scalar(text(row.get('publication_short')))}",
            f"abstract: {yaml_scalar(text(row.get('abstract')))}",
            f"summary: {yaml_scalar(text(row.get('summary')))}",
        ]
    )
    if tags:
        lines.append("tags:")
        lines.extend(f"  - {yaml_scalar(tag)}" for tag in tags)
    else:
        lines.append("tags: []")
    lines.extend(
        [
            f"featured: {yaml_scalar(parse_bool(row.get('featured')))}",
            f"url_pdf: {yaml_scalar(url_pdf)}",
            f"url_code: {yaml_scalar(text(row.get('url_code')))}",
            f"url_dataset: {yaml_scalar(text(row.get('url_dataset')))}",
            f"url_poster: {yaml_scalar(text(row.get('url_poster')))}",
            f"url_project: {yaml_scalar(text(row.get('url_project')))}",
            f"url_slides: {yaml_scalar(text(row.get('url_slides')))}",
            f"url_source: {yaml_scalar(text(row.get('url_source')))}",
            f"url_video: {yaml_scalar(text(row.get('url_video')))}",
            "image:",
            f"  caption: {yaml_scalar(text(row.get('caption')))}",
            "  focal_point: Center",
            "  preview_only: false",
            "projects: []",
            f"slides: {yaml_scalar(text(row.get('slides')))}",
            "---",
            "",
        ]
    )
    body = body_from_markdown(row, markdown_dir)
    if body:
        lines.extend([body, ""])
    return folder, "\n".join(lines)


def generate(input_path: Path, output_dir: Path, pictures_dir: Path, markdown_dir: Path, pdfs_dir: Path) -> int:
    rows = read_rows(input_path)
    count = 0
    for row in rows:
        rendered = render_publication(row, markdown_dir)
        if rendered is None:
            continue
        folder, content = rendered
        publication_dir = output_dir / folder
        publication_dir.mkdir(parents=True, exist_ok=True)
        (publication_dir / "index.md").write_text(content, encoding="utf-8")
        picture = text(row.get("picture"))
        if picture:
            source = pictures_dir / picture
            if source.is_file():
                shutil.copyfile(source, publication_dir / f"featured{source.suffix.lower()}")
            else:
                print(f"WARNING: missing publication picture: {source}")
        copy_optional_file(text(row.get("pdf_file")), pdfs_dir, publication_dir)
        count += 1
    print(f"Generated {count} publication pages in {output_dir}")
    return count


def main() -> None:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--input", type=Path, default=DEFAULT_INPUT)
    parser.add_argument("--pictures", type=Path, default=DEFAULT_PICTURES)
    parser.add_argument("--markdown", type=Path, default=DEFAULT_MARKDOWN)
    parser.add_argument("--pdfs", type=Path, default=DEFAULT_PDFS)
    parser.add_argument("--output", type=Path, default=DEFAULT_OUTPUT)
    args = parser.parse_args()
    generate(args.input, args.output, args.pictures, args.markdown, args.pdfs)


if __name__ == "__main__":
    main()
